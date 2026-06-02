#!/usr/bin/env python3
"""
export_config.py — Excel → Lua 配置文件导出工具

将 design/*.xlsx 中的策划配置表导出为 Lua 配置文件。
支持单 key / 多 key 索引、json 复杂类型。

用法:
  python tools/export_config.py               # 导出 design/ 下所有 .xlsx
  python tools/export_config.py design/item.xlsx  # 导出单个文件

Excel 格式约定（4 行表头）:
  Row 1: 字段说明（中文注释）
  Row 2: 字段名（用作 Lua key）
  Row 3: 字段类型 (int|float|bool|string|json)
  Row 4: 主键标记 (#1, #2, ... 空表示非 key)
  Row 5+: 数据行
"""

import json
import os
import sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    print("[ERROR] 需要 openpyxl: pip install openpyxl")
    sys.exit(1)

# 路径
TOOL_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(TOOL_DIR, ".."))
DESIGN_DIR = os.path.join(PROJECT_ROOT, "design")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "server", "module", "config")
CLIENT_OUTPUT_DIR = os.path.join(PROJECT_ROOT, "client", "config")

HEADER_ROWS = 4  # 表头行数（desc, name, type, key）


# ======== Lua 值序列化 ========

def to_lua_literal(obj):
    """将 Python 对象转换为 Lua 字面量字符串"""
    if obj is None:
        return "nil"
    elif isinstance(obj, bool):
        return "true" if obj else "false"
    elif isinstance(obj, int):
        return str(obj)
    elif isinstance(obj, float):
        if obj == int(obj):
            return str(int(obj))
        return str(obj)
    elif isinstance(obj, str):
        escaped = (
            obj.replace("\\", "\\\\")
            .replace('"', '\\"')
            .replace("\n", "\\n")
            .replace("\r", "\\r")
            .replace("\t", "\\t")
        )
        return f'"{escaped}"'
    elif isinstance(obj, (list, tuple)):
        items = [to_lua_literal(v) for v in obj]
        return "{" + ", ".join(items) + "}"
    elif isinstance(obj, dict):
        items = []
        for k, v in obj.items():
            key_part = k if isinstance(k, str) and k.isidentifier() else f'["{to_lua_key(k)}"]'
            items.append(f"{key_part} = {to_lua_literal(v)}")
        return "{" + ", ".join(items) + "}"
    return "nil"


def to_lua_key(val):
    """将 Python 值转换为 Lua table key 字符串"""
    if isinstance(val, str):
        escaped = val.replace("\\", "\\\\").replace('"', '\\"')
        return escaped
    return str(val)


def flatten_keys(d):
    """将嵌套 OrderedDict 展平为 [(key_path_list, value_str), ...]"""
    result = []
    for k, v in d.items():
        if isinstance(v, OrderedDict):
            sub = flatten_keys(v)
            for path, val_str in sub:
                result.append(([k] + path, val_str))
        else:
            result.append(([k], v))
    return result


def serialize_lua_fields(field_names, field_types, row_values):
    """将一行数据序列化为 Lua 表字段列表。

    返回: 字符串列表，如 ['id = 1001', 'name = "HP药水"']
    """
    parts = []
    for i, name in enumerate(field_names):
        if i >= len(row_values):
            continue
        val = row_values[i]
        if val is None:
            continue
        typ = field_types[i] if i < len(field_types) else "string"

        if typ == "json" and isinstance(val, str):
            # JSON 字符串解析后转 Lua table
            try:
                parsed = json.loads(val)
                lua_val = to_lua_literal(parsed)
            except json.JSONDecodeError:
                lua_val = to_lua_literal(val)
        else:
            lua_val = to_lua_literal(val)

        parts.append(f"{name} = {lua_val}")
    return parts


# ======== Excel 解析 ========

def parse_sheet(ws, filename=""):
    """解析单个 sheet，返回结构化数据。"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < HEADER_ROWS + 1:
        return None  # 没有数据行

    # 读取表头
    descs = [str(c) if c is not None else "" for c in rows[0]]  # noqa: 字段说明（仅用于注释）
    names = [str(c).strip() if c is not None else "" for c in rows[1]]
    types = [str(c).strip() if c is not None else "" for c in rows[2]]
    keys = [str(c).strip() if c is not None else "" for c in rows[3]]

    # 清理：过滤掉完全空白的列
    valid_cols = []
    for i, name in enumerate(names):
        if name:
            valid_cols.append(i)

    field_names = [names[i] for i in valid_cols]
    field_types = [types[i] for i in valid_cols]
    field_keys = [keys[i] for i in valid_cols]
    field_descs = [descs[i] for i in valid_cols]

    # 找到 key 列
    key_indices = []  # [(col_index, level)]
    for i, k in enumerate(field_keys):
        if k.startswith("#"):
            try:
                level = int(k[1:])
                key_indices.append((i, level))
            except ValueError:
                pass

    # 按 level 排序
    key_indices.sort(key=lambda x: x[1])

    # 解析数据行
    data_rows = []
    warnings = []

    def warn(row_num, field_name, field_desc, msg):
        loc = f"{filename} / {ws.title} / 第{row_num}行 / {field_name}({field_desc})"
        warnings.append(f"  [WARN] {loc} : {msg}")

    for row_idx in range(HEADER_ROWS, len(rows)):
        row = rows[row_idx]
        values = [row[i] if i < len(row) else None for i in valid_cols]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue

        data_row_num = row_idx + 1  # Excel 行号（1-based）
        converted = []
        for i, v in enumerate(values):
            if v is None:
                converted.append(None)
                continue
            typ = field_types[i]
            fname = field_names[i]
            fdesc = field_descs[i]

            if typ == "int":
                try:
                    converted.append(int(float(str(v))))
                except (ValueError, TypeError):
                    warn(data_row_num, fname, fdesc, f"整数格式错误: '{v}' 无法转换为整数")
                    converted.append(None)
            elif typ == "float":
                try:
                    converted.append(float(str(v)))
                except (ValueError, TypeError):
                    warn(data_row_num, fname, fdesc, f"浮点数格式错误: '{v}' 无法转换为浮点数")
                    converted.append(None)
            elif typ == "bool":
                if isinstance(v, bool):
                    converted.append(v)
                else:
                    s = str(v).strip().lower()
                    converted.append(s in ("true", "1", "yes"))
                    if s not in ("true", "false", "1", "0", "yes", "no"):
                        warn(data_row_num, fname, fdesc, f"布尔值格式错误: '{v}' 不是有效的布尔值，已按 false 处理")
            elif typ == "json":
                if isinstance(v, str):
                    s = v.strip()
                    if s:
                        try:
                            json.loads(s)
                        except json.JSONDecodeError as e:
                            warn(data_row_num, fname, fdesc, f"JSON 解析失败: {e}")
                    converted.append(s if s else None)
                else:
                    converted.append(str(v) if v is not None else None)
            else:  # string
                converted.append(str(v).strip() if v is not None else None)

        data_rows.append(converted)

    return {
        "sheet_name": ws.title,
        "field_names": field_names,
        "field_types": field_types,
        "key_indices": key_indices,
        "data_rows": data_rows,
        "warnings": warnings,
    }


# ======== Lua 生成 ========

def set_nested(d, keys, val_str):
    """向嵌套 OrderedDict 中按 key 路径插入值。"""
    if len(keys) == 1:
        d[keys[0]] = val_str
    else:
        k = keys[0]
        if k not in d:
            d[k] = OrderedDict()
        set_nested(d[k], keys[1:], val_str)


def generate_lua(sheet_data):
    """从解析后的 sheet 数据生成 Lua 代码。"""
    field_names = sheet_data["field_names"]
    field_types = sheet_data["field_types"]
    key_indices = sheet_data["key_indices"]
    data_rows = sheet_data["data_rows"]

    lines = []
    lines.append("-- auto-generated by tools/export_config.py -- do not edit")
    lines.append("")

    # 判断是否有 key
    if key_indices:
        data = OrderedDict()

        for row in data_rows:
            key_parts = []
            for idx, _ in key_indices:
                val = row[idx] if idx < len(row) else None
                key_parts.append(val)

            # 构建值字段
            val_parts = serialize_lua_fields(field_names, field_types, row)
            val_str = "{" + ", ".join(val_parts) + "}"

            if len(key_indices) == 1:
                k = key_parts[0]
                if k is not None:
                    data[k] = val_str
            else:
                set_nested(data, key_parts, val_str)

        # 输出
        lines.append("local M = {}")
        lines.append("")
        if len(key_indices) == 1:
            for k, v in data.items():
                k_str = to_lua_literal(k)
                lines.append(f"M[{k_str}] = {v}")
        else:
            seen_mid = set()
            for key_path, val_str in flatten_keys(data):
                # 初始化中间表（如 M[1001] = {}）
                parent = "M"
                for i, k in enumerate(key_path[:-1]):
                    parent += f"[{to_lua_literal(k)}]"
                    if parent not in seen_mid:
                        seen_mid.add(parent)
                        lines.append(f"{parent} = {{}}")
                # 最终赋值
                expr = parent + f"[{to_lua_literal(key_path[-1])}]"
                lines.append(f"{expr} = {val_str}")

    else:
        # 无 key：生成数组表
        lines.append("local M = {}")
        lines.append("")
        for _, row in enumerate(data_rows, 1):
            val_parts = serialize_lua_fields(field_names, field_types, row)
            lines.append(f"M[{_}] = " + "{" + ", ".join(val_parts) + "}")

    lines.append("")
    lines.append("return M")
    lines.append("")
    return "\n".join(lines)


# ======== 导出逻辑 ========

def export_file(xlsx_path):
    """导出单个 .xlsx 文件。"""
    filename = os.path.basename(xlsx_path)
    mod_name = filename.replace(".xlsx", "").replace(".XLSX", "")

    print(f"\n{'='*50}")
    print(f"  [export] {filename}")
    print(f"{'='*50}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("_"):
            print(f"  [SKIP] sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        data = parse_sheet(ws, filename)
        if data is None:
            print(f"  [SKIP] {sheet_name}: 无数据行")
            continue

        # 显示警告
        warnings = data.get("warnings", [])
        if warnings:
            print(f"  [{len(warnings)} warnings in sheet: {sheet_name}]")
            for w in warnings:
                print(w)

        # 生成 Lua
        try:
            lua_code = generate_lua(data)
        except Exception as e:
            print(f"  [ERROR] {sheet_name} Lua 生成失败: {e}")
            continue

        # 写入服务端
        server_path = os.path.join(OUTPUT_DIR, f"{sheet_name}.lua")
        with open(server_path, "w", encoding="utf-8") as f:
            f.write(lua_code)
        print(f"  [OK] → {os.path.relpath(server_path, PROJECT_ROOT)}")

        # 预留客户端输出
        client_path = os.path.join(CLIENT_OUTPUT_DIR, f"{sheet_name}.lua")
        with open(client_path, "w", encoding="utf-8") as f:
            f.write(lua_code)
        print(f"  [OK] → {os.path.relpath(client_path, PROJECT_ROOT)}")

    wb.close()


def export_all():
    """导出 design/ 下所有 .xlsx。"""
    if not os.path.isdir(DESIGN_DIR):
        print(f"[ERROR] 目录不存在: {DESIGN_DIR}")
        return

    for fname in sorted(os.listdir(DESIGN_DIR)):
        if fname.startswith("_"):
            continue
        if fname.endswith(".xlsx") or fname.endswith(".XLSX"):
            export_file(os.path.join(DESIGN_DIR, fname))


# ======== 创建示例 Excel ========

def create_sample_excel():
    """生成 design/item.xlsx 示例文件。"""
    os.makedirs(DESIGN_DIR, exist_ok=True)
    path = os.path.join(DESIGN_DIR, "item.xlsx")
    if os.path.exists(path):
        print(f"[skip] {path} already exists")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "item"

    # Row 1: 字段说明
    ws.append(["ID", "名称", "类型", "子类型", "品质", "可堆叠", "最大堆叠", "出售价格",
               "使用效果", "装备槽位", "基础属性", "描述"])
    # Row 2: 字段名
    ws.append(["id", "name", "type", "sub_type", "quality", "stackable", "max_stack",
               "sell_price", "use_effect", "equip_slot", "attributes", "description"])
    # Row 3: 字段类型
    ws.append(["int", "string", "string", "string", "int", "bool", "int",
               "int", "json", "string", "json", "string"])
    # Row 4: 主键标记
    ws.append(["#1", "", "", "", "", "", "", "", "", "", "", ""])

    # 数据行 - 对应 item_template.lua 的现有数据
    ws.append([1001, "小瓶HP药水", "consumable", "hp_potion", 1, True, 99, 10,
               '{"hp":50}', "", "", "恢复50点生命值"])
    ws.append([1002, "小瓶MP药水", "consumable", "mp_potion", 1, True, 99, 10,
               '{"mp":50}', "", "", "恢复50点魔法值"])
    ws.append([2001, "铁矿", "material", "ore", 1, True, 999, 5,
               "", "", "", "常见的铁矿石"])
    ws.append([3001, "铁剑", "equipment", "sword", 1, False, 1, 50,
               "", "weapon", '{"attack":10}', "一把普通的铁剑"])
    ws.append([3002, "铁甲", "equipment", "chest", 1, False, 1, 80,
               "", "chest", '{"defense":15}', "一件普通的铁甲"])

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"[create] {path}")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(CLIENT_OUTPUT_DIR, exist_ok=True)

    if len(sys.argv) > 1 and sys.argv[1] != "--all":
        export_file(sys.argv[1])
    else:
        print("=" * 50)
        print("  Excel → Lua Config Export")
        print("=" * 50)
        create_sample_excel()
        export_all()
        print("=" * 50)
        print("  Done!")
        print("=" * 50)


if __name__ == "__main__":
    main()
